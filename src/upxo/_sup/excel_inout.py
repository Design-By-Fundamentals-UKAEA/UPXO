import numpy as np
from openpyxl import Workbook, load_workbook

def write_to_excel(wb, dictionary, filename):
    """Export or convert to te to excel."""

    for i, (sheet_name, variable) in enumerate(dictionary.items()):
        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        if isinstance(variable, np.ndarray):
            if variable.ndim == 1:
                for r in range(variable.shape[0]):
                    ws.cell(row=r+1, column=1, value=variable[r])
            else:
                for r in range(variable.shape[0]):
                    for c in range(variable.shape[1]):
                        ws.cell(row=r+1, column=c+1, value=variable[r, c])
        else:
            ws.cell(row=1, column=1, value=variable)

    wb.save(filename)

def read_from_excel(filename, sheet_names):
    """Load or import from excel."""
    wb = load_workbook(filename)
    data_from_sheets = {}
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        if ws.max_row == 1 and ws.max_column == 1:
            data_from_sheets[sheet_name] = ws.cell(row=1, column=1).value
        else:
            if ws.max_column == 1:
                data = np.array([ws.cell(row=r+1, column=1).value for r in range(ws.max_row)])
            else:
                data = np.array([[ws.cell(row=r+1, column=c+1).value for c in range(ws.max_column)] for r in range(ws.max_row)])
            data_from_sheets[sheet_name] = data
    return data_from_sheets
